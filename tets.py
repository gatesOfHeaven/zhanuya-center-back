# from asyncio import run
# from redis.asyncio import from_url
# from datetime import timedelta


# async def main():
#     redis = from_url('redis://localhost')
#     # await redis.set('my-val', 'value of it', timedelta(seconds = 15))
#     val: bytes = await redis.get('email:meteorite.medik@gmail.com')
#     print(val.decode())


# run(main())

import sqlite3
from openpyxl import Workbook

def export_sqlite_to_excel(db_path, output_path):
    # Connect to the SQLite database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Create a new Excel workbook
    workbook = Workbook()
    
    # Fetch all table names from the database
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()
    
    # Iterate over all tables and add each to a new sheet
    for table_name in tables:
        table_name = table_name[0]  # Extract table name from tuple
        cursor.execute(f"SELECT * FROM {table_name}")
        rows = cursor.fetchall()
        column_names = [description[0] for description in cursor.description]

        # Create a new sheet or use the default one for the first table
        if len(workbook.sheetnames) == 1 and workbook.active.title == "Sheet":
            sheet = workbook.active
            sheet.title = table_name
        else:
            sheet = workbook.create_sheet(title=table_name)

        # Write the column headers
        sheet.append(column_names)

        # Write all rows
        for row in rows:
            sheet.append(row)
    
    # Save the workbook
    workbook.save(output_path)
    print(f"Data exported successfully to {output_path}")

    # Close the database connection
    conn.close()

# Usage example
sqlite_db_path = "test.db"
output_excel_path = "output.xlsx"
export_sqlite_to_excel(sqlite_db_path, output_excel_path)
